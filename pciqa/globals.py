from openpyxl import load_workbook
from pathlib import Path
from tkinter import Tk
from tkinter.filedialog import askdirectory
import sys




def set_workdir():
    """ Set the directory that holds the evidence to be tested.

        This program is designed to be packaged as an exe with PyInstaller. When run as an exe, a Tkinter window lets the
        user select the memo directory but for testing/development use a static test directory.

        This function will also parse a command line argument for a directory name. So you can run the program like so:

        python matlock testdata/

        This only for testing is not an option using the EXE

        Returns:
        workdir: Path object to the directory where memo files will be read.
    """

    if getattr(sys, 'frozen', False):  # Program is running from an exe
        Tk().withdraw()  # Launch TKinter. Don't need full GUI, so keep root window from appearing
        return Path(askdirectory()) # Open a file-browser to let user select the memo folder
    else: # Program is being called directly, outside of an exe
        return Path(f"{sys.argv[-1]}")

WORKDIR = set_workdir()



'''
def read_vars_file():
	wb = load_workbook('openpyxl.xlsx')
	for c1, c8 in zip(sheet.iter_rows(min_col=1, max_col=1), sheet.iter_rows(min_col=8, max_col=8)):
'''


req_checkbox_types = {
	'☐☐☐☐☐': 'empty',
    '☒☐☐☐☐': 'In Place',
    '☒☐☐☐': 'In Place', # For 3.2.1 - 3.2.3, the N/A checkbox is blanked out
    '☐☒☐☐☐': 'In Place w/CCW',
    '☐☒☐☐': 'In Place w/CCW', # For 3.2.1 - 3.2.3, the N/A checkbox is blanked out
    '☐☐☒☐☐': 'N/A',
    '☐☐☐☒☐': 'Not Tested',
    '☐☐☒☐': 'Not Tested', # For 3.2.1 - 3.2.3, the N/A checkbox is blanked out
    '☐☐☐☐☒': 'Not in Place',
    '☐☐☐☒': 'Not in Place', # For 3.2.1 - 3.2.3, the N/A checkbox is blanked out
    '☒☒☐☐☐': 'multiple',
    '☒☐☒☐☐': 'multiple',
    '☒☐☐☒☐': 'multiple',
    '☒☐☐☐☒': 'multiple',
    '☐☒☒☐☐': 'multiple',
    '☐☒☐☒☐': 'multiple',
    '☐☒☐☐☒': 'multiple',
    '☐☐☒☒☐': 'multiple',
    '☐☐☒☐☒': 'multiple',
    '☐☐☐☒☒': 'multiple'
}